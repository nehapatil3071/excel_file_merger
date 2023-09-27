from flask import Flask, request, render_template, send_file
import os
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

@app.route("/", methods=["GET", "POST"])
def upload_excel():
    if request.method == "POST":
        uploaded_files = request.files.getlist("excel_files")
        merged_workbook = Workbook()

        for file in uploaded_files:
            if file.filename != "":
                try:
                    workbook = load_workbook(file)
                    for sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        merged_workbook.create_sheet(title=sheet.title)

                        for row in sheet.iter_rows():
                            merged_sheet = merged_workbook[sheet_name]
                            merged_sheet.append([cell.value for cell in row])

                except Exception as e:
                    return f"Error processing {file.filename}: {str(e)}"

        merged_filename = "merged_excel.xlsx"
        merged_workbook.remove(merged_workbook.active)  # Remove default sheet
        merged_workbook.save(merged_filename)

        return send_file(
            merged_filename,
            as_attachment=True,
            download_name=merged_filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    return render_template("upload_excel.html")

if __name__ == "__main__":
    app.run(debug=True)
